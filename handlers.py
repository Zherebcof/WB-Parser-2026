# handlers.py — Файл с логикой (Мозг бота)
# Здесь мы решаем, что отвечать на команды и нажатия кнопок

import os  # 📂 Работа с файловой системой (удаление файлов)
import asyncio  # ⏱ Асинхронность (паузы)
import logging  # 📝 Логирование ошибок
from aiogram import Router, F, types  # 🤖 Инструменты Aiogram (Роутер, Фильтры, Типы)
from aiogram.filters import Command  # 🔗 Фильтр для команд (например, /start)
from aiogram.fsm.context import FSMContext  # 🧠 Управление памятью (машина состояний)
from aiogram.fsm.state import State, StatesGroup  # 🚦 Состояния (шаги диалога)
from aiogram.types import (ReplyKeyboardMarkup, KeyboardButton, FSInputFile,
                           InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery)
import pandas as pd  # 🐼 Библиотека для работы с таблицами данных

# 👇 Инструменты для красивого Excel (Админ-панель)
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

import database  # 🗄️ Наша база данных

# Попытка подключить парсер. Если файла нет — бот не упадет, а просто скажет "ошибка".
try:
    from wb_parser import get_wb_data, save_to_excel
except ImportError:
    get_wb_data = None
    save_to_excel = None

# ==========================================
# ⚙️ НАСТРОЙКИ
# ==========================================
router = Router()  # 📡 Создаем Роутер (он ловит сообщения и направляет в нужную функцию)

# 📝 Белый список (кто может пользоваться ботом)
ALLOWED_USERS = [
    812808959,  # Анатолий
]

# 👑 Список Админов (у кого есть доступ к /admin)
ADMINS = [
    812808959,  # Анатолий
]


# 🚦 Машина состояний (FSM)
# Это как светофор: бот запоминает, на каком шаге мы находимся
class SearchStates(StatesGroup):
    waiting_for_item = State()  # Ждем название товара
    waiting_for_broadcast = State()  # Ждем текст для рассылки


# ==========================================
# ⌨️ КЛАВИАТУРЫ (КНОПКИ)
# ==========================================

# Главное меню (внизу экрана)
main_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🔎 Поиск товара")],  # Большая кнопка поиска
        [KeyboardButton(text="ℹ️ Помощь")]  # Кнопка помощи
    ],
    resize_keyboard=True  # Делаем кнопки компактными
)

# Админ-меню (под сообщением)
admin_keyboard = InlineKeyboardMarkup(inline_keyboard=[
    [InlineKeyboardButton(text="📊 Статистика", callback_data="admin_stats")],
    [InlineKeyboardButton(text="📥 Скачать базу", callback_data="admin_excel")],
    [InlineKeyboardButton(text="📢 Рассылка", callback_data="admin_broadcast")]
])


# ==========================================
# 👮‍♂️ ПОМОЩНИК (ПРОВЕРКА ДОСТУПА)
# ==========================================
def is_user_allowed(user_id):
    """Проверяет, есть ли пользователь в белом списке"""
    return user_id in ALLOWED_USERS


# ==========================================
# 🎮 ХЭНДЛЕРЫ (ОБРАБОТЧИКИ)
# ==========================================

# 1. КОМАНДА /START
@router.message(Command("start"))
async def cmd_start(message: types.Message):
    # Проверка прав
    if not is_user_allowed(message.from_user.id):
        await message.answer("⛔ Доступ к боту запрещен.")
        return

    # Регистрируем пользователя в базе данных (если его там нет)
    await database.add_user(message.from_user.id, message.from_user.username, message.from_user.full_name)

    # Приветствуем и показываем кнопки
    await message.answer("Привет! 👋 Я готов искать товары.\nНажми кнопку ниже ⬇️", reply_markup=main_keyboard)


# 2. НАЖАТИЕ КНОПКИ "🔎 ПОИСК ТОВАРА"
@router.message(F.text == "🔎 Поиск товара")
async def start_search_scenario(message: types.Message, state: FSMContext):
    if not is_user_allowed(message.from_user.id): return

    await message.answer("Введите название товара (например: 'Кроссовки Nike'):")
    # Переводим бота в режим "Жду название товара"
    await state.set_state(SearchStates.waiting_for_item)


# 3. ПОЛУЧЕНИЕ НАЗВАНИЯ ТОВАРА (САМ ПОИСК)
@router.message(SearchStates.waiting_for_item)
async def process_item_search(message: types.Message, state: FSMContext):
    if not is_user_allowed(message.from_user.id): return
    user_text = message.text  # То, что написал юзер

    # Если юзер передумал и ввел команду
    if user_text.startswith('/'):
        await state.clear()
        if user_text == "/admin": await open_admin_panel(message)
        return

    # Если нажал "Помощь" вместо товара
    if user_text == "ℹ️ Помощь":
        await state.clear()
        await cmd_help(message, state)
        return

    # ✅ НАЧИНАЕМ ПОИСК
    await message.answer(f"🔎 Ищу '{user_text}'... Подождите пару секунд.")

    try:
        # 1. Записываем запрос в историю (в базу)
        await database.add_search_query(message.from_user.id, user_text)

        # 2. Запускаем Парсер (функция из файла wb_parser.py)
        products = await get_wb_data(user_text)

        # Если товаров нет
        if not products:
            await message.answer("Ничего не нашел по этому запросу 😔")
            return

        # 3. Сохраняем в Excel (функция из wb_parser.py)
        filename = f"wb_{user_text.replace(' ', '_')}.xlsx"
        save_to_excel(products, filename)

        # 4. Отправляем файл пользователю
        excel_file = FSInputFile(filename)
        await message.answer_document(excel_file, caption=f"📊 Отчет по запросу: {user_text}")

        # 5. Удаляем файл с сервера (чтобы не засорять память)
        try:
            os.remove(filename)
        except:
            pass

        # 6. Показываем 5 карточек с фото (для красоты)
        for product in products[:5]:
            caption = f"🔹 <b>{product['name']}</b>\n💰 {product['price']}\n🔗 <a href='{product['link']}'>Ссылка на товар</a>"

            if product.get('image'):
                await message.answer_photo(product['image'], caption=caption, parse_mode="HTML")
            else:
                await message.answer(caption, parse_mode="HTML")
            await asyncio.sleep(0.5)  # Небольшая пауза, чтобы сообщения шли по порядку

        # Завершаем сценарий
        await state.clear()
        await message.answer("Готово! ✅ Можем искать снова.", reply_markup=main_keyboard)

    # 🆘 ОБРАБОТКА ОШИБОК (ТОТ САМЫЙ БЛОК)
    except Exception as e:
        logging.error(f"🔥 КРИТИЧЕСКАЯ ОШИБКА: {e}")
        await message.answer("😔 Произошла техническая ошибка. Разработчик уже уведомлен!")

        # Шлем отчет Админу (Тебе)
        try:
            admin_id = ADMINS[0]
            error_text = f"🆘 <b>АВАРИЯ!</b>\nЮзер: {message.from_user.id}\nОшибка: {str(e)[:500]}"
            await message.bot.send_message(chat_id=admin_id, text=error_text, parse_mode="HTML")
        except:
            pass
        await state.clear()


# 4. КОМАНДА ПОМОЩЬ
@router.message(F.text == "ℹ️ Помощь")
async def cmd_help(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("👨‍💻 Разработчик: Анатолий Жеребцов\nБот парсит Wildberries в реальном времени.")


# ==========================================
# 👑 АДМИН-ПАНЕЛЬ
# ==========================================

@router.message(Command("admin"))
async def open_admin_panel(message: types.Message):
    if message.from_user.id not in ADMINS: return
    await message.answer("👑 Панель Администратора", reply_markup=admin_keyboard)


# Кнопка "Статистика"
@router.callback_query(F.data == "admin_stats")
async def show_stats(callback: CallbackQuery):
    if callback.from_user.id not in ADMINS: return
    report = await database.get_stats()  # Берем цифры из базы
    await callback.answer()
    await callback.message.edit_text(report, reply_markup=admin_keyboard, parse_mode="HTML")


# Кнопка "Скачать базу" (С красивым дизайном)
@router.callback_query(F.data == "admin_excel")
async def export_excel_db(callback: CallbackQuery):
    if callback.from_user.id not in ADMINS: return
    await callback.answer("🎨 Генерирую отчет...")

    users = await database.get_all_users()
    if not users:
        await callback.message.answer("База пуста")
        return

    # Создаем Excel
    filename = "users_base.xlsx"
    df = pd.DataFrame(users, columns=['ID', 'Никнейм', 'Имя', 'Дата регистрации'])
    df.to_excel(filename, index=False)

    # 🎨 КРАСИМ ТАБЛИЦУ
    wb = load_workbook(filename)
    ws = wb.active
    # Настраиваем ширину и цвета (как делали раньше)
    for column in ws.columns:
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = 25
        for cell in column:
            cell.alignment = Alignment(horizontal='center')
    # Синяя шапка
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    wb.save(filename)

    # Отправляем
    file = FSInputFile(filename)
    await callback.message.answer_document(file, caption="📂 База пользователей")
    os.remove(filename)


# Кнопка "Рассылка"
@router.callback_query(F.data == "admin_broadcast")
async def start_broadcast(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMINS: return
    await callback.answer()
    await callback.message.answer("📢 Введите текст (или фото) для рассылки всем:")
    await state.set_state(SearchStates.waiting_for_broadcast)


# Процесс рассылки
@router.message(SearchStates.waiting_for_broadcast)
async def process_broadcast(message: types.Message, state: FSMContext):
    if message.from_user.id not in ADMINS: return

    # Отмена
    if message.text and message.text.lower() == 'отмена':
        await state.clear()
        await message.answer("Отменено ❌")
        return

    # Получаем всех пользователей
    users_ids = await database.get_all_users_ids()
    count = 0
    await message.answer(f"🚀 Начинаю рассылку на {len(users_ids)} человек...")

    # Рассылаем
    for user_id in users_ids:
        try:
            # copy_to отправляет копию сообщения (текст, фото, видео - всё что угодно)
            await message.copy_to(chat_id=user_id)
            count += 1
            await asyncio.sleep(0.1)  # Анти-спам
        except:
            pass

    await message.answer(f"✅ Рассылка завершена! Доставлено: {count}")
    await state.clear()