# wb_parser.py — Модуль парсинга (Руки бота)
# Здесь происходит магия сбора данных с сайта Wildberries

import asyncio  # ⏱ Для пауз и асинхронности
import pandas as pd  # 🐼 Для создания таблиц
from playwright.async_api import async_playwright  # 🎭 Браузер для парсинга

# 🎨 Инструменты для дизайна Excel (Маляр)
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ⚙️ Селекторы (адреса элементов на сайте)
# Если WB поменяет дизайн сайта, менять настройки нужно только здесь
SELECTORS = {
    "card_wrapper": ".product-card__wrapper",  # Карточка товара
    "link_tag": ".product-card__link",  # Ссылка
    "price_new": ".price__lower-price",  # Новая цена
    "price_old": ".price__wrap del",  # Старая цена (зачеркнутая)
    "image": "img"  # Картинка
}


async def get_wb_data(query):
    """
    Главная функция парсинга.
    Открывает браузер, ищет товар `query` и возвращает список словарей.
    """
    async with async_playwright() as p:
        # Запускаем скрытый браузер (headless=True)
        # args=['...'] нужны, чтобы сайт не понял, что мы робот
        browser = await p.chromium.launch(
            headless=True,
            args=['--disable-blink-features=AutomationControlled', '--no-sandbox']
        )
        # Создаем "контекст" (как отдельное окно инкогнито)
        context = await browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            viewport={'width': 1920, 'height': 1080}
        )
        page = await context.new_page()

        # 🕵️‍♂️ СТЕЛС-РЕЖИМ (Маскировка)
        # Подменяем свойства браузера, чтобы WB думал, что мы человек
        await page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

        # Формируем ссылку поиска
        url = f"https://www.wildberries.ru/catalog/0/search.aspx?search={query}"

        try:
            print(f"🌍 Иду на сайт: {url}")
            await page.goto(url, wait_until='domcontentloaded', timeout=40000)

            # Ждем появления карточек товаров
            await page.wait_for_selector(SELECTORS["card_wrapper"], timeout=15000)
            product_cards = await page.query_selector_all(SELECTORS["card_wrapper"])

            results = []

            # 🔄 Проходим по первым 15 товарам
            for card in product_cards[:15]:
                try:
                    # 1. Ссылка и Название
                    link_tag = await card.query_selector(SELECTORS["link_tag"])
                    if not link_tag: continue

                    name = await link_tag.get_attribute('aria-label')
                    href = await link_tag.get_attribute('href')

                    # 2. Цены
                    price_new_tag = await card.query_selector(SELECTORS["price_new"])
                    price_old_tag = await card.query_selector(SELECTORS["price_old"])

                    # Пытаемся найти хоть какую-то цену
                    if price_new_tag:
                        price = await price_new_tag.inner_text()
                    elif price_old_tag:
                        price = await price_old_tag.inner_text()
                    else:
                        price = "Нет цены"

                    old_price = await price_old_tag.inner_text() if price_old_tag else None

                    # 3. Картинка
                    img_tag = await card.query_selector(SELECTORS["image"])
                    img_link = await img_tag.get_attribute('src') if img_tag else None

                    # Исправляем ссылку (добавляем https)
                    if img_link and img_link.startswith('//'):
                        img_link = 'https:' + img_link

                    # Добавляем товар в список
                    results.append({
                        'name': name,
                        'price': price,
                        'old_price': old_price,
                        'link': href,
                        'image': img_link
                    })
                except Exception as e:
                    # Если одна карточка сломалась, пропускаем её, но не ломаем весь парсер
                    print(f"⚠️ Ошибка чтения карточки: {e}")
                    continue

            return results

        except Exception as e:
            print(f"🔥 Авария парсера! Ошибка: {e}")
            return []

        finally:
            await browser.close()


def save_to_excel(data, filename):
    """
    Сохраняет данные в Excel и делает 'Евроремонт' (Дизайн).
    """
    # 1. СОЗДАЕМ ТАБЛИЦУ (PANDAS)
    df = pd.DataFrame(data)

    # Переименовываем колонки на Русский язык
    df.rename(columns={
        'name': 'Название товара',
        'price': 'Цена',
        'old_price': 'Старая цена',
        'link': 'Ссылка',
        'image': 'Фото'
    }, inplace=True)

    # Сохраняем "сырой" файл
    df.to_excel(filename, index=False)

    # ==========================================
    # 🎨 2. НАВОДИМ КРАСОТУ (OpenPyXL)
    # ==========================================
    wb = load_workbook(filename)
    ws = wb.active

    # Настройки для Шапки
    header_fill = PatternFill(start_color="104E8B", end_color="104E8B", fill_type="solid")  # Синий фон
    header_font = Font(bold=True, color="FFFFFF", size=11)  # Белый текст
    center_align = Alignment(horizontal='center', vertical='center')  # По центру

    # А. Красим Шапку (первая строка)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Б. Настраиваем ширину колонок
    for col in ws.columns:
        column_letter = col[0].column_letter
        header_value = col[0].value
        max_length = 0

        for cell in col:
            # Название товара — выравниваем слева
            if header_value == "Название товара":
                if cell.row > 1: cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.column_dimensions[column_letter].width = 40

                # Ссылка — делаем кликабельной
            elif header_value == "Ссылка":
                ws.column_dimensions[column_letter].width = 50
                if cell.value and cell.row > 1:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")

            # Остальное — авто-ширина
            else:
                cell.alignment = center_align
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except:
                    pass

        # Если это не Название и не Ссылка — ставим ширину по содержимому
        if header_value not in ["Название товара", "Ссылка"]:
            ws.column_dimensions[column_letter].width = (max_length + 5)

    wb.save(filename)
    print(f"🎨 Excel украшен: {filename}")